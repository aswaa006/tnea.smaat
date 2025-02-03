import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def process_csv(file_path, output_file):
    # Read CSV file
    df = pd.read_csv(file_path)
    
    # Ensure required columns exist
    required_columns = {"7.5% reservation", "general", "total", "oc", "bc", "bcm", "mbc", "sc", "sca", "st", 
                        "oc_7.5", "bc_7.5", "bcm_7.5", "mbc_7.5", "sc_7.5", "sca_7.5", "st_7.5"}
    
    if not required_columns.issubset(df.columns):
        missing_cols = required_columns - set(df.columns)
        raise ValueError(f"Missing required columns: {missing_cols}")
    
    # First computed sum column
    df['general+7.5% reservation'] = df['7.5% reservation'] + df['general']
    
    # Second computed sum column (General Sum)
    df['General Sum'] = df[['oc', 'bc', 'bcm', 'mbc', 'sc', 'sca', 'st']].sum(axis=1)
    
    # Third computed sum column (7.5% Reservation Sum)
    df['7.5% Reservation Sum'] = df[['oc_7.5', 'bc_7.5', 'bcm_7.5', 'mbc_7.5', 'sc_7.5', 'sca_7.5', 'st_7.5']].sum(axis=1)
    
    # Save to Excel with color formatting
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())
    
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for index, row in df.iterrows():
        ws.append(row.tolist())
        row_idx = index + 2  # Excel rows start from 1, header is row 1
        
        # Apply color formatting
       
    
    wb.save(output_file)
    print(f"Processed file saved as {output_file}")

# Example usage
process_csv("3_2_1_govtreserved (8).csv", "4_3_2_1_govtreserved.xlsx")
